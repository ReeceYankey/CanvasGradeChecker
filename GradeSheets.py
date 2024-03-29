import openpyxl as xl
from copy import copy
from collections import Counter

"""
data:
category section
assignment types are primary sort

"""



class WeightedSheetHandler:
    def __init__(self, ws):
        self.ws = ws

    def is_category_header(self, row):
        try:
            return self.ws.cell(row=row, column=3).value == 'Due Date' and \
                   '=B' in self.ws.cell(row=row, column=2).value
        except (AttributeError, ValueError):
            return False

    def is_category_totals_row(self, row):
        try:
            return self.ws.cell(row=row, column=3).value == 'Total:' and \
                   '=B' in self.ws.cell(row=row, column=2).value
        except (AttributeError, ValueError):
            return False

    def is_row_insert_advisory(self, row):
        return self.ws.cell(row=row, column=2).value is not None and \
               '*If you need to add a row' in self.ws.cell(row=row, column=2).value

    def get_category_header_rows(self):
        """get the row numbers of the header rows"""
        return [r for r in range(36, self.ws.max_row+1) if self.is_category_header(r)]

    def get_category_totals_rows(self):
        """get the row numbers of the totals rows"""
        return [r for r in range(37, self.ws.max_row+1) if self.is_category_totals_row(r)]

    def add_row(self, row):
        self.ws.insert_rows(row)
        self.update_pointers()
        self.style_added_cells(row)

    def style_added_cells(self, row):
        """copies the styling from row-1 to row"""

        for i in [2, 8, 15]:
            for col in range(i, i + 4):
                self.ws.cell(row=row, column=col)._style = copy(self.ws.cell(row=row - 1, column=col)._style)

    def update_pointers(self):
        # ----update pointers at top (C16:D25)-----
        ws = self.ws

        ending_rows = self.get_category_totals_rows()
        header_rows = self.get_category_header_rows()

        # loop over grade breakdown totals
        for r in range(16, 26):
            for c in range(3, 5):

                cell = ws.cell(row=r, column=c)
                ending_row_index = (r - 16) // 3 # which set of horizontal category boxes
                cell.value = cell.value[:2] + str(ending_rows[ending_row_index]) # '=column'  +  'ending row index'

        # find grade entry start and end rows for purpose of finding bounds of pointers
        grade_entries_starts = [i+1 for i in header_rows]
        grade_entries_ends = [i-3 for i in ending_rows]

        # replace pointers at category totals rows
        for i, row in enumerate(ending_rows):
            POINTS_EARNED_COLS = [4, 10, 17] # TODO switch to D, J, Q
            POINTS_POSSIBLE_COLS = [5, 11, 18]
            # replace points earned sums
            for col in POINTS_EARNED_COLS:
                cell = ws.cell(row=row, column=col)
                column_letter = xl.utils.get_column_letter(col)
                cell.value = '=SUM({}:{})'.format(column_letter + str(grade_entries_starts[i]),
                                                  column_letter + str(grade_entries_ends[i]))
            # replace points possible sums
            for col in POINTS_POSSIBLE_COLS:
                cell = ws.cell(row=row, column=col)
                column_letter_1 = xl.utils.get_column_letter(col - 1)
                column_letter_2 = xl.utils.get_column_letter(col)
                cell.value = '=SUMIF({}:{},">=0",{}:{})'.format(column_letter_1 + str(grade_entries_starts[i]),
                                                                column_letter_1 + str(grade_entries_ends[i]),
                                                                column_letter_2 + str(grade_entries_starts[i]),
                                                                column_letter_2 + str(grade_entries_ends[i]))
        # remove extra unnecessary pointers to the right of category 10
        ws['J{}'.format(ending_rows[-1])].value = None
        ws['K{}'.format(ending_rows[-1])].value = None
        ws['Q{}'.format(ending_rows[-1])].value = None
        ws['R{}'.format(ending_rows[-1])].value = None

    def unmerge_ending_cells(self):
        totals_rows = self.get_category_totals_rows()
        for c in [2, 8, 15]: # leftmost col in a category
            for r in totals_rows:
                try:
                    self.ws.unmerge_cells(start_row=r - 2, start_column=c, end_row=r - 1, end_column=c + 3)
                except ValueError as e:
                    print(e)

    def merge_ending_cells(self):
        totals_rows = self.get_category_totals_rows()
        for c in [2, 8, 15]: # leftmost col in a category
            for r in totals_rows:
                self.ws.merge_cells(start_row=r - 2, start_column=c, end_row=r - 1, end_column=c + 3)

    def update(self, table):
        # TODO: fix potential conflict: table can be either list or pandas.core.series.Series, force type
        # merged cells cannot be modified, so all cells that are shifted must be unmerged then remerged at the end
        self.unmerge_ending_cells()
        # set categories and weightings
        # TODO look into weightings
        categories = list(Counter(table['type']))
        for r in range(len(categories)):
            self.ws.cell(row=r + 16, column=2).value = categories[r] # TODO 16 is grade breakdown start

        # add assignments to corresponding sections
        self.add_all_assignments(table)

        self.merge_ending_cells()

    def add_all_assignments(self, table):
        print('adding all assignments')
        ws = self.ws
        # add assignments to corresponding sections
        # starts at top left, goes down until finds an unused category section, then returns back up and to the right

        header_rows = self.get_category_header_rows()
        for category_row in header_rows:
            for category_col in [2, 8, 15]:
                curr_row = category_row
                # if sheet_row >= ws.max_row:
                #     print("reached max row")
                #     continue

                cell = ws.cell(row=curr_row, column=category_col)
                if cell.value is None:
                    continue
                category_title = ws[cell.value[1:]].value  # gets the value of the cell that the cell points to
                if category_title is None:
                    continue

                # t = table['type'].tolist()
                # assert category_title in table['type'].tolist()

                # add data to section
                curr_row += 1
                for i, assign_type in enumerate(table['type']):
                    if assign_type == category_title:
                        if self.is_row_insert_advisory(curr_row):
                            self.add_row(curr_row)
                        self.fill_grade_entry(curr_row, category_col, table['name'][i], table['date'][i], table['score'][i], table['max_score'][i])
                        curr_row += 1

                # clear old data from section
                # print(type(ws.cell(row=curr_row, column=category_col).value))
                while not self.is_row_insert_advisory(curr_row):
                    self.fill_grade_entry(curr_row, category_col, None, None, None, None)
                    curr_row += 1

    def fill_grade_entry(self, sheet_row, sheet_col, name, date, score, max_score):
        self.ws.cell(row=sheet_row, column=sheet_col).value = name
        self.ws.cell(row=sheet_row, column=sheet_col + 1).value = date
        self.ws.cell(row=sheet_row, column=sheet_col + 2).value = score
        self.ws.cell(row=sheet_row, column=sheet_col + 3).value = max_score


class PointSheetHandler:
    def __init__(self, ws):
        self.ws = ws

    def is_totals_row(self, row):
        return self.ws.cell(row=row, column=2).value == 'Total'

    def get_totals_row(self):
        ws = self.ws

        for r in range(37, ws.max_row + 1):
            if self.is_totals_row(r):
                return r

    def add_row(self, row, count=1):
        self.ws.insert_rows(row, amount=count)
        self.update_pointers()
        self.style_added_cells(row)

    def style_added_cells(self, row):
        """copies the styling from row-1 to row"""
        for col in range(1, 8):
            self.ws.cell(row=row, column=col)._style = copy(self.ws.cell(row=row - 1, column=col)._style)

    def update_pointers(self):
        # ----update pointers at top-----
        totals_row = self.get_totals_row()

        cell = self.ws['K8']
        cell.value = '={} / {} * 100'.format('D' + str(totals_row),
                                             'G' + str(totals_row))
        cell = self.ws['K9']
        cell.value = '={}'.format('D' + str(totals_row))

        grade_entries_start = 16  # the start for the area that holds assignment data
        grade_entries_end = totals_row - 1  # the end for the area that holds assignment data

        cell = self.ws.cell(row=totals_row, column=4)
        cell.value = '=SUM({}:{})'.format('D' + str(grade_entries_start),
                                          'D' + str(grade_entries_end))
        cell = self.ws.cell(row=totals_row, column=7)
        cell.value = '=SUMIF({}:{},">=0",{}:{})'.format('D' + str(grade_entries_start),
                                                        'D' + str(grade_entries_end),
                                                        'G' + str(grade_entries_start),
                                                        'G' + str(grade_entries_end))
        cell = self.ws['R16']
        cell.value = '=SUM({}:{})'.format('G' + str(grade_entries_start),
                                          'G' + str(grade_entries_end))

    def unmerge_ending_cells(self):
        ws = self.ws

        totals_row = self.get_totals_row()
        ws.unmerge_cells(start_row=totals_row + 1, start_column=1, end_row=totals_row + 2, end_column=7)

    def merge_ending_cells(self):
        ws = self.ws

        totals_row = self.get_totals_row()
        print('merging', totals_row + 1, 1, totals_row + 2, 7)
        ws.merge_cells(start_row=totals_row + 1, start_column=1, end_row=totals_row + 2, end_column=7)

    def update(self, table):
        self.unmerge_ending_cells()
        self.add_all_assignments(table)
        self.merge_ending_cells()

    def add_all_assignments(self, table):
        # add data to section
        sheet_row = 16
        for i in range(len(table['name'])):
            if self.is_totals_row(sheet_row):
                self.add_row(sheet_row)
            self.fill_grade_entry(sheet_row, 2, table['name'][i], table['date'][i], table['score'][i], table['max_score'][i])
            sheet_row += 1

        # clear old data from section
        # print(type(ws.cell(row=sheet_row, column=sheet_col).value))
        while not self.is_totals_row(sheet_row):
            self.fill_grade_entry(sheet_row, 2, None, None, None, None)
            sheet_row += 1
        
    def fill_grade_entry(self, sheet_row, sheet_col, name, date, score, max_score):
        self.ws.cell(row=sheet_row, column=sheet_col).value = name
        self.ws.cell(row=sheet_row, column=sheet_col + 1).value = date
        self.ws.cell(row=sheet_row, column=sheet_col + 2).value = score
        self.ws.cell(row=sheet_row, column=sheet_col + 5).value = max_score
